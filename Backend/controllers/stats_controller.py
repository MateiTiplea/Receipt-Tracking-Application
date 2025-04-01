from fastapi import APIRouter, HTTPException, Response, Query
from Backend.firebase_config import client
from google.cloud import bigquery
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
import matplotlib.pyplot as plt

stats_router = APIRouter(prefix="/api/v1/stats", tags=["Statistics"])


@stats_router.get("/top-stores")
def get_top_stores():
    query = """
        SELECT
          JSON_VALUE(data, '$.store_name') AS store_name,
          COUNT(*) AS total_bonuri,
          SUM(CAST(JSON_VALUE(data, '$.total_amount') AS FLOAT64)) AS total_cheltuit
        FROM `receipt-tracking-application.receipts_info.receipts_raw_latest`
        GROUP BY store_name
        ORDER BY total_cheltuit DESC
        LIMIT 5
    """
    try:
        result = client.query(query).result()
        return [
            {
                "store_name": row.store_name,
                "bonuri": row.total_bonuri,
                "total_cheltuit": row.total_cheltuit
            } for row in result
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare BigQuery: {e}")


@stats_router.get("/CRUD-operations-distribution")
def get_operation_counts():
    query = """
    SELECT 
      operation,
      COUNT(*) AS count
    FROM `receipt-tracking-application.receipts_info.receipts_raw_changelog`
    GROUP BY operation
    ORDER BY count DESC
    """

    try:
        results = client.query(query).result()
        return [{"operation": row.operation, "count": row.count} for row in results]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare BigQuery: {e}")


@stats_router.get("/users-spendings")
def get_top_users():
    query = """
    SELECT 
      JSON_VALUE(data, '$.user_uid') AS user_uid,
      SUM(CAST(JSON_VALUE(data, '$.total_amount') AS FLOAT64)) AS total_spent
    FROM `receipt-tracking-application.receipts_info.receipts_raw_latest`
    GROUP BY user_uid
    ORDER BY total_spent DESC
    """

    try:
        results = client.query(query).result()
        rows = [
            {
                "user_uid": row.user_uid,
                "total_spent": row.total_spent
            }
            for row in results
        ]
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare BigQuery: {e}")


@stats_router.get("/change-log-info")
def get_change_log():
    query = """
    SELECT *
    FROM `receipt-tracking-application.receipts_info.receipts_raw_changelog`
    ORDER BY timestamp DESC
    """

    try:
        results = client.query(query).result()
        rows = [dict(row.items()) for row in results]
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare BigQuery: {e}")


@stats_router.get("/export-excel")
def export_excel_for_user(uid: str = Query(..., description="UID-ul utilizatorului")):
    # Query 1: Get receipt details
    receipts_query = """
    SELECT
      JSON_VALUE(data, '$.store_name') AS store_name,
      CAST(JSON_VALUE(data, '$.total_amount') AS FLOAT64) AS total_amount,
      JSON_VALUE(data, '$.date') AS date
    FROM `receipt-tracking-application.receipts_info.receipts_raw_latest`
    WHERE JSON_VALUE(data, '$.user_uid') = @uid
    ORDER BY date DESC
    """

    # Query 2: Get category spending data
    categories_query = """
    SELECT
      category,
      SUM(CAST(JSON_VALUE(data, '$.total_amount') AS FLOAT64)) AS total_spent
    FROM `receipt-tracking-application.receipts_info.receipts_raw_latest`,
    UNNEST(JSON_VALUE_ARRAY(data, '$.categories')) AS category
    WHERE JSON_VALUE(data, '$.user_uid') = @uid
    GROUP BY category
    ORDER BY total_spent DESC
    """

    # Execute first query for receipts
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("uid", "STRING", uid)]
    )
    receipts_job = client.query(receipts_query, job_config=job_config)
    receipts_results = receipts_job.result()
    receipts_rows = list(receipts_results)

    if not receipts_rows:
        raise HTTPException(status_code=404, detail="Nu s-au găsit bonuri pentru acest utilizator.")

    # Execute second query for categories
    categories_job = client.query(categories_query, job_config=job_config)
    categories_results = categories_job.result()
    categories_rows = list(categories_results)

    # Create DataFrames
    df_receipts = pd.DataFrame([dict(r.items()) for r in receipts_rows])
    df_receipts["date"] = pd.to_datetime(df_receipts["date"])
    total_spent = df_receipts["total_amount"].sum()

    # Get top stores data (from original first method)
    top_stores = df_receipts["store_name"].value_counts().reset_index()
    top_stores.columns = ["store_name", "num_receipts"]

    if categories_rows:
        df_categories = pd.DataFrame([dict(r.items()) for r in categories_rows])
    else:
        # Create empty DataFrame if no categories found
        df_categories = pd.DataFrame(columns=["category", "total_spent"])

    # Create Excel workbook
    wb = Workbook()

    # First tab: Receipts list
    ws_bonuri = wb.active
    ws_bonuri.title = "Bonuri"

    # Add receipt data
    for r in dataframe_to_rows(df_receipts, index=False, header=True):
        ws_bonuri.append(r)

    # Format date column
    for row in ws_bonuri.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "DD-MM-YYYY"

    # Second tab: Store frequency with bar chart
    ws_stores = wb.create_sheet(title="Magazine")
    ws_stores.append(["Total cheltuit (RON)", round(total_spent, 2)])
    ws_stores.append([])
    ws_stores.append(["Top magazine", "Nr. bonuri"])

    # Add store frequency data
    for index, row in top_stores.iterrows():
        ws_stores.append([row["store_name"], row["num_receipts"]])

    # Create bar chart for store frequency
    plt.figure(figsize=(6, 4))
    plt.bar(top_stores["store_name"], top_stores["num_receipts"])
    plt.xticks(rotation=45, ha='right')
    plt.title("Cele mai frecventate magazine")
    plt.tight_layout()

    stores_img_buffer = BytesIO()
    plt.savefig(stores_img_buffer, format='png')
    stores_img_buffer.seek(0)

    stores_img = XLImage(stores_img_buffer)
    ws_stores.add_image(stores_img, "E2")

    # Third tab: Categories with pie chart
    ws_categories = wb.create_sheet(title="Categorii")
    ws_categories.append(["Categorie", "Suma totală (RON)"])

    # Add category data
    for index, row in df_categories.iterrows():
        ws_categories.append([row["category"], round(row["total_spent"], 2)])

    # Create pie chart for categories if data exists
    if not df_categories.empty:
        plt.figure(figsize=(6, 6))
        plt.pie(df_categories["total_spent"], labels=df_categories["category"],
                autopct="%1.1f%%", startangle=140)
        plt.title("Distribuția cheltuielilor pe categorii")
        plt.tight_layout()

        categories_img_buffer = BytesIO()
        plt.savefig(categories_img_buffer, format='png')
        categories_img_buffer.seek(0)

        categories_img = XLImage(categories_img_buffer)
        ws_categories.add_image(categories_img, "D2")

    # Save and return Excel file
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.getvalue()
    filename = f"raport_{uid}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )